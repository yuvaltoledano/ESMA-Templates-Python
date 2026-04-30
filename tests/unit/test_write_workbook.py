"""Unit tests for the workbook writer's encoding rules.

The most important test here pins the **date encoding contract**:
Polars `pl.Date` columns must round-trip through openpyxl as Excel
serial integers (matching what R's openxlsx writes), NOT as
`datetime.date` objects. This test is the gate that should fail loudly
if a future refactor switches to openpyxl's native date support
(which would break parity vs the staged R fixture).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import polars as pl

from esma_milan.io_layer.write_workbook import (
    _date_to_excel_serial,
    write_pipeline_workbook,
)

# ---------------------------------------------------------------------------
# _date_to_excel_serial
# ---------------------------------------------------------------------------


def test_date_to_excel_serial_anchor() -> None:
    """The synthetic fixture's origination_date for ORIG_001 is
    2019-03-15, written as Excel serial 43539. Pin this end-to-end:
    if my serial conversion drifts, this will fail with a clear
    integer mismatch."""
    assert _date_to_excel_serial(date(2019, 3, 15)) == 43539


def test_date_to_excel_serial_round_trip_against_parse() -> None:
    """Round-trip: parse_iso_or_excel_date(serial) -> date, and
    _date_to_excel_serial(date) -> serial, must be inverses for every
    value present in the synthetic and real fixtures."""
    # Sample of serials drawn from the synthetic fixture's loans/
    # properties date columns.
    from esma_milan.io_layer.date_parsing import parse_iso_or_excel_date_values

    serials = [43539, 54497, 45473, 45550, 45306, 43983, 54940, 54321]
    parsed = parse_iso_or_excel_date_values(serials, "test")
    re_serialised = [_date_to_excel_serial(p) for p in parsed if p is not None]
    assert re_serialised == serials


def test_date_to_excel_serial_at_excel_epoch() -> None:
    """Boundary: 1899-12-30 -> 0; 1899-12-31 -> 1; 1900-01-01 -> 2."""
    assert _date_to_excel_serial(date(1899, 12, 30)) == 0
    assert _date_to_excel_serial(date(1899, 12, 31)) == 1
    assert _date_to_excel_serial(date(1900, 1, 1)) == 2


# ---------------------------------------------------------------------------
# write_pipeline_workbook: encoding round-trips
# ---------------------------------------------------------------------------


def test_writer_encodes_pl_date_as_excel_serial_int(tmp_path: Path) -> None:
    """The critical contract: Polars pl.Date column -> openpyxl reads
    back as int. NOT datetime.date. R's openxlsx produces int on read,
    so this is the byte-equal-via-roundtrip target."""
    df = pl.DataFrame(
        {
            "calc_loan_id": pl.Series(["L1", "L2"], dtype=pl.String),
            "origination_date": pl.Series(
                [date(2019, 3, 15), date(2020, 6, 1)], dtype=pl.Date
            ),
            "collateral_group_id": pl.Series([1, 2], dtype=pl.Int64),
        }
    )
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(out, populated_sheets={"Cleaned ESMA loans": df})

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Cleaned ESMA loans"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    # Header
    assert rows[0] == ("calc_loan_id", "origination_date", "collateral_group_id")
    # Row 1: date column is int 43539, NOT a datetime object
    assert rows[1] == ("L1", 43539, 1)
    assert isinstance(rows[1][1], int)
    assert not isinstance(rows[1][1], bool)
    # Row 2 likewise
    assert rows[2] == ("L2", 43983, 2)


def test_writer_encodes_null_date_as_empty_cell(tmp_path: Path) -> None:
    """Null Date values must produce empty Excel cells. openpyxl's
    write-only mode drops trailing None cells from the row, so reading
    back returns a shortened tuple. The parity diff library handles
    this correctly via `c_idx < len(row) else None` - see
    parity/diff.py::_diff_sheet."""
    df = pl.DataFrame(
        {
            "calc_loan_id": pl.Series(["L1"], dtype=pl.String),
            "date_last_in_arrears": pl.Series([None], dtype=pl.Date),
        }
    )
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(out, populated_sheets={"Cleaned ESMA loans": df})

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Cleaned ESMA loans"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    # openpyxl drops the trailing None -> row is length 1, not 2.
    assert rows[1] == ("L1",)


def test_writer_preserves_embedded_none_cells(tmp_path: Path) -> None:
    """Embedded None cells (between non-None values) ARE preserved by
    openpyxl write_only - only TRAILING Nones are dropped. Pin this so
    we know which Nones survive when designing parity expectations."""
    df = pl.DataFrame(
        {
            "id": pl.Series(["L1"], dtype=pl.String),
            "missing_middle": pl.Series([None], dtype=pl.String),
            "tail": pl.Series(["X"], dtype=pl.String),
        }
    )
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(out, populated_sheets={"Cleaned ESMA loans": df})

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Cleaned ESMA loans"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1] == ("L1", None, "X")


def test_writer_passes_through_non_date_dtypes(tmp_path: Path) -> None:
    """Schema dtypes other than pl.Date round-trip natively through
    openpyxl: Int64 -> int, Float64 -> float (or int if whole), Boolean
    -> bool, String -> str, null -> None."""
    df = pl.DataFrame(
        {
            "id": pl.Series(["L1", "L2"], dtype=pl.String),
            "balance_int": pl.Series([220000, 350000], dtype=pl.Int64),
            "rate_float": pl.Series([2.5, 3.1], dtype=pl.Float64),
            "active": pl.Series([True, False], dtype=pl.Boolean),
            "optional": pl.Series([None, "value"], dtype=pl.String),
        }
    )
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(out, populated_sheets={"Cleaned ESMA loans": df})

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Cleaned ESMA loans"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "balance_int", "rate_float", "active", "optional")
    # Row 1 ends in None -> openpyxl drops it; row is length 4.
    assert rows[1] == ("L1", 220000, 2.5, True)
    assert rows[2] == ("L2", 350000, 3.1, False, "value")
    # Spot-check types match what the synthetic fixture surfaces:
    assert isinstance(rows[1][1], int) and not isinstance(rows[1][1], bool)
    assert isinstance(rows[1][2], float)
    assert isinstance(rows[1][3], bool)


def test_writer_handles_mix_of_date_and_non_date_columns(tmp_path: Path) -> None:
    """A realistic Cleaned ESMA loans-style frame with multiple date
    columns interspersed with non-date columns. All dates encoded;
    everything else passes through."""
    df = pl.DataFrame(
        {
            "id": pl.Series(["L1"], dtype=pl.String),
            "origination_date": pl.Series([date(2019, 3, 15)], dtype=pl.Date),
            "balance": pl.Series([180000], dtype=pl.Int64),
            "maturity_date": pl.Series([date(2049, 3, 15)], dtype=pl.Date),
            "active": pl.Series([True], dtype=pl.Boolean),
            "date_last_in_arrears": pl.Series([None], dtype=pl.Date),
        }
    )
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(out, populated_sheets={"Cleaned ESMA loans": df})

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Cleaned ESMA loans"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    # Trailing None (date_last_in_arrears) dropped by openpyxl; non-trailing
    # date columns surface as ints.
    assert rows[1] == (
        "L1",
        43539,
        180000,
        _date_to_excel_serial(date(2049, 3, 15)),
        True,
    )


def test_writer_unpopulated_sheets_stay_empty(tmp_path: Path) -> None:
    """Populating one sheet leaves the other 9 empty. Sheet order
    matches OUTPUT_SHEET_ORDER."""
    from esma_milan.config import OUTPUT_SHEET_ORDER

    df = pl.DataFrame({"x": pl.Series([1], dtype=pl.Int64)})
    out = tmp_path / "test.xlsx"
    write_pipeline_workbook(
        out, populated_sheets={"Group classifications": df}
    )

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert wb.sheetnames == list(OUTPUT_SHEET_ORDER)
    for name in wb.sheetnames:
        ws = wb[name]
        ws.reset_dimensions()
        rows = list(ws.iter_rows(values_only=True))
        if name == "Group classifications":
            assert rows == [("x",), (1,)]
        else:
            assert rows == []
