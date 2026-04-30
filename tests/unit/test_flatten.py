"""Tests for esma_milan.pipeline.flatten.

B-3 covers `select_main_property`: 5-key priority sort with the
mandatory deliberate-tie fixtures (lex-smallest calc_property_id wins
when keys 1-4 all tie) plus determinism anchors (reversed property_id
order; shuffled row order) and NA-sort-key behaviour for each of the
four primary keys.
"""

from __future__ import annotations

import warnings
from collections.abc import Sequence
from datetime import date
from pathlib import Path

import polars as pl

from esma_milan.pipeline.flatten import select_main_property


def _props(
    *,
    group_ids: Sequence[int],
    property_ids: Sequence[str],
    occupancy_types: Sequence[str | None],
    valuation_methods: Sequence[str | None],
    valuation_dates: Sequence[date | None],
) -> pl.DataFrame:
    """Properties frame post-select_valuation_fields, minimal columns."""
    return pl.DataFrame(
        {
            "collateral_group_id": pl.Series(list(group_ids), dtype=pl.Int64),
            "calc_property_id": pl.Series(list(property_ids), dtype=pl.String),
            "occupancy_type": pl.Series(list(occupancy_types), dtype=pl.String),
            "valuation_method": pl.Series(list(valuation_methods), dtype=pl.String),
            "valuation_date": pl.Series(list(valuation_dates), dtype=pl.Date),
        }
    )


def _unique_vals(
    group_ids: Sequence[int],
    property_ids: Sequence[str],
    values: Sequence[float | None],
) -> pl.DataFrame:
    return pl.DataFrame(
        {
            "collateral_group_id": pl.Series(list(group_ids), dtype=pl.Int64),
            "calc_property_id": pl.Series(list(property_ids), dtype=pl.String),
            "property_value": pl.Series(list(values), dtype=pl.Float64),
        }
    )


# ---------------------------------------------------------------------------
# Key 1: highest property_value wins
# ---------------------------------------------------------------------------


def test_highest_property_value_wins() -> None:
    """3 properties in one group, distinct values -> highest wins."""
    props = _props(
        group_ids=[1, 1, 1],
        property_ids=["P_A", "P_B", "P_C"],
        occupancy_types=["FOWN", "FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 3,
    )
    vals = _unique_vals([1, 1, 1], ["P_A", "P_B", "P_C"], [200000.0, 500000.0, 300000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


# ---------------------------------------------------------------------------
# Key 2: owner-occupied (FOWN) wins on value tie
# ---------------------------------------------------------------------------


def test_value_tie_breaks_to_owner_occupied() -> None:
    """Same value -> FOWN preferred over non-FOWN."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["TLET", "FOWN"],
        valuation_methods=["FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


# ---------------------------------------------------------------------------
# Key 3: full-inspection method (FIEI/FOEI) wins on (value, FOWN) tie
# ---------------------------------------------------------------------------


def test_value_and_fown_tie_breaks_to_full_inspection() -> None:
    """Same value, both FOWN -> full-inspection method wins."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=["DRVB", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


def test_full_inspection_includes_foei_alongside_fiei() -> None:
    """FOEI is also full-inspection per VALUATION_METHODS_FULL_INSPECTION."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=["DRVB", "FOEI"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


# ---------------------------------------------------------------------------
# Key 4: most-recent valuation_date wins on (value, FOWN, method) tie
# ---------------------------------------------------------------------------


def test_value_fown_method_tie_breaks_to_most_recent_date() -> None:
    """Same value/FOWN/FIEI -> most-recent date wins."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI"],
        valuation_dates=[date(2020, 1, 1), date(2024, 1, 1)],
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


# ---------------------------------------------------------------------------
# Key 5: lex tie-break (THE MANDATORY DELIBERATE-TIE FIXTURE)
# ---------------------------------------------------------------------------


def _deliberate_tie_fixture() -> tuple[pl.DataFrame, pl.DataFrame]:
    """Three properties tied on every key 1-4 - only calc_property_id
    distinguishes them. Lex-smallest must win."""
    props = _props(
        group_ids=[1, 1, 1],
        property_ids=["P_C", "P_A", "P_B"],  # deliberately scrambled
        occupancy_types=["FOWN", "FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 3,
    )
    vals = _unique_vals(
        [1, 1, 1], ["P_C", "P_A", "P_B"], [500000.0, 500000.0, 500000.0]
    )
    return props, vals


def test_lex_tie_break_picks_smallest_calc_property_id() -> None:
    """All four primary keys tie -> lex-smallest calc_property_id wins."""
    props, vals = _deliberate_tie_fixture()
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_A")


def test_lex_tie_break_invariant_to_reversed_property_id_input_order() -> None:
    """Reverse the row order of the deliberate-tie fixture - lex-smallest
    must still win. Catches "lex-smallest implemented as lex-largest"
    where the implementation might accidentally pick the LAST sorted row."""
    props, vals = _deliberate_tie_fixture()
    out_normal = select_main_property(props, vals)
    out_reversed = select_main_property(props.reverse(), vals.reverse())
    assert out_normal.row(0) == (1, "P_A")
    assert out_reversed.row(0) == (1, "P_A")
    assert out_normal.equals(out_reversed)


def test_lex_tie_break_invariant_to_arbitrary_input_row_shuffle() -> None:
    """Shuffle input rows in a non-trivial permutation. Mirrors B-2's
    determinism anchor: the answer must be identical regardless of
    input row order, since the sort is supposed to be a pure function
    of row content. Catches sort-stability or hidden-ordering leakage
    that the reversed-input test alone would miss."""
    props, vals = _deliberate_tie_fixture()

    # An arbitrary non-trivial permutation: rotate by 1 (move first row
    # to the end). Use Polars row index + sort to materialise the
    # permutation deterministically.
    rotated_props = pl.concat([props.slice(1, 2), props.slice(0, 1)])
    rotated_vals = pl.concat([vals.slice(1, 2), vals.slice(0, 1)])

    out_normal = select_main_property(props, vals)
    out_rotated = select_main_property(rotated_props, rotated_vals)
    assert out_normal.row(0) == (1, "P_A")
    assert out_rotated.row(0) == (1, "P_A")
    assert out_normal.equals(out_rotated)


# ---------------------------------------------------------------------------
# NA-sort-key behaviour (Polars vs R sort-order parity)
# ---------------------------------------------------------------------------


def test_null_property_value_sorts_after_non_null() -> None:
    """NA property_value -> sorts after non-NA. With nulls_last=True
    descending, the null-value row loses despite no other distinguishing
    features. Catches a future Polars change where descending sort
    moves nulls first."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    # P_A null value, P_B has 100k. Even though P_A's lex is smaller,
    # the null sorts last on key 1 -> P_B wins.
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [None, 100000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


def test_null_occupancy_type_sorts_after_fown_and_non_fown() -> None:
    """Per user direction: occupancy_type == "FOWN" produces null when
    the input is null, and Polars vs R may sort that null differently
    under desc(). With nulls_last=True the null sorts AFTER both True
    (FOWN) and False (non-FOWN) values.

    Fixture: two rows, all keys except occupancy_type tied. Row B has
    occupancy_type=TLET (False), Row A has occupancy_type=None (Null).
    P_A is lex-smaller, so if lex tie-break decided, P_A would win. But
    the boolean key 2 fires first: False > Null -> P_B wins.
    """
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=[None, "TLET"],
        valuation_methods=["FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


def test_null_valuation_method_sorts_after_inspection_and_non_inspection() -> None:
    """Same shape as above for key 3 (`valuation_method ∈ {FIEI, FOEI}`).

    Fixture: two rows tied on keys 1-2, differ on key 3. Row B has
    valuation_method=DRVB (False on inspection check), Row A has
    valuation_method=None (Null). P_A is lex-smaller. Boolean key 3
    fires: False > Null -> P_B wins.
    """
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=[None, "DRVB"],
        valuation_dates=[date(2024, 1, 1)] * 2,
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


def test_null_valuation_date_sorts_after_non_null_dates() -> None:
    """Key 4 (valuation_date) - null sorts after non-null with nulls_last=True
    descending. Non-null date wins."""
    props = _props(
        group_ids=[1, 1],
        property_ids=["P_A", "P_B"],
        occupancy_types=["FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI"],
        valuation_dates=[None, date(2020, 1, 1)],
    )
    vals = _unique_vals([1, 1], ["P_A", "P_B"], [500000.0, 500000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_B")


# ---------------------------------------------------------------------------
# Multi-group + edge cases
# ---------------------------------------------------------------------------


def test_multi_group_each_independently() -> None:
    """Multiple groups in one input - each group's main property is
    selected independently."""
    props = _props(
        group_ids=[1, 1, 2, 2, 3],
        property_ids=["P_A", "P_B", "P_X", "P_Y", "P_Z"],
        occupancy_types=["FOWN", "FOWN", "FOWN", "FOWN", "FOWN"],
        valuation_methods=["FIEI", "FIEI", "FIEI", "FIEI", "FIEI"],
        valuation_dates=[date(2024, 1, 1)] * 5,
    )
    vals = _unique_vals(
        [1, 1, 2, 2, 3],
        ["P_A", "P_B", "P_X", "P_Y", "P_Z"],
        [200000.0, 500000.0, 300000.0, 100000.0, 400000.0],
    )
    out = select_main_property(props, vals)
    by_gid = {row[0]: row[1] for row in out.iter_rows()}
    assert by_gid == {1: "P_B", 2: "P_X", 3: "P_Z"}


def test_single_property_per_group_picks_that_property() -> None:
    """Group with one property -> that one wins trivially."""
    props = _props(
        group_ids=[1],
        property_ids=["P_ONLY"],
        occupancy_types=["FOWN"],
        valuation_methods=["FIEI"],
        valuation_dates=[date(2024, 1, 1)],
    )
    vals = _unique_vals([1], ["P_ONLY"], [100000.0])
    out = select_main_property(props, vals)
    assert out.row(0) == (1, "P_ONLY")


def test_output_one_row_per_group_sorted_by_group_id() -> None:
    """Output is exactly one row per collateral_group_id, sorted by
    group_id ascending."""
    props = _props(
        group_ids=[3, 1, 2, 3, 1],
        property_ids=["P_3a", "P_1a", "P_2a", "P_3b", "P_1b"],
        occupancy_types=["FOWN"] * 5,
        valuation_methods=["FIEI"] * 5,
        valuation_dates=[date(2024, 1, 1)] * 5,
    )
    vals = _unique_vals(
        [3, 1, 2, 3, 1],
        ["P_3a", "P_1a", "P_2a", "P_3b", "P_1b"],
        [100.0, 200.0, 300.0, 400.0, 500.0],
    )
    out = select_main_property(props, vals)
    assert out["collateral_group_id"].to_list() == [1, 2, 3]


def test_output_schema() -> None:
    """Schema check: collateral_group_id Int64, main_property_id String."""
    props = _props(
        group_ids=[1],
        property_ids=["P_A"],
        occupancy_types=["FOWN"],
        valuation_methods=["FIEI"],
        valuation_dates=[date(2024, 1, 1)],
    )
    vals = _unique_vals([1], ["P_A"], [100000.0])
    out = select_main_property(props, vals)
    assert out.schema["collateral_group_id"] == pl.Int64
    assert out.schema["main_property_id"] == pl.String
    assert out.columns == ["collateral_group_id", "main_property_id"]


# ---------------------------------------------------------------------------
# B-4: flatten_loan_collateral orchestrator
# ---------------------------------------------------------------------------

import math  # noqa: E402

import pytest  # noqa: E402

from esma_milan.pipeline.classification import (  # noqa: E402
    TYPE_1,
    TYPE_2,
    TYPE_3,
    TYPE_4,
    TYPE_5,
)
from esma_milan.pipeline.flatten import flatten_loan_collateral  # noqa: E402


def _loans_enriched(rows: Sequence[tuple[str, int, float]]) -> pl.DataFrame:
    """Build loans_enriched with (calc_loan_id, collateral_group_id, CPB)."""
    return pl.DataFrame(
        {
            "calc_loan_id": pl.Series([r[0] for r in rows], dtype=pl.String),
            "collateral_group_id": pl.Series([r[1] for r in rows], dtype=pl.Int64),
            "current_principal_balance": pl.Series(
                [r[2] for r in rows], dtype=pl.Float64
            ),
        }
    )


def _properties_enriched(
    rows: Sequence[tuple[str, str, int, str, float | None]],
    *,
    occupancy_type: str = "FOWN",
    current_method: str = "FIEI",
    original_method: str = "FIEI",
    current_date: date = date(2024, 1, 1),
    original_date: date = date(2020, 1, 1),
    original_amount: float = 100.0,
) -> pl.DataFrame:
    """Build properties_enriched with (loan_ref, property_id, group_id,
    current_method, current_amount). Other valuation columns default to
    sane values; tests can override via kwargs (or build by hand)."""
    return pl.DataFrame(
        {
            "underlying_exposure_identifier": pl.Series(
                [r[0] for r in rows], dtype=pl.String
            ),
            "calc_property_id": pl.Series([r[1] for r in rows], dtype=pl.String),
            "collateral_group_id": pl.Series([r[2] for r in rows], dtype=pl.Int64),
            "current_valuation_method": pl.Series(
                [r[3] for r in rows], dtype=pl.String
            ),
            "current_valuation_amount": pl.Series(
                [r[4] for r in rows], dtype=pl.Float64
            ),
            "current_valuation_date": pl.Series(
                [current_date] * len(rows), dtype=pl.Date
            ),
            "original_valuation_method": pl.Series(
                [original_method] * len(rows), dtype=pl.String
            ),
            "original_valuation_amount": pl.Series(
                [original_amount] * len(rows), dtype=pl.Float64
            ),
            "original_valuation_date": pl.Series(
                [original_date] * len(rows), dtype=pl.Date
            ),
            "occupancy_type": pl.Series([occupancy_type] * len(rows), dtype=pl.String),
        }
    )


def _classification(rows: Sequence[tuple[int, str]]) -> pl.DataFrame:
    """Minimal group_classification frame with (gid, structure_type)."""
    return pl.DataFrame(
        {
            "collateral_group_id": pl.Series(
                [r[0] for r in rows], dtype=pl.Int64
            ),
            "structure_type": pl.Series([r[1] for r in rows], dtype=pl.String),
        }
    )


# ---------------------------------------------------------------------------
# Type 1-5: structural surface
# ---------------------------------------------------------------------------


def test_flatten_type1_single_loan_single_property() -> None:
    """1 loan -> 1 property: pro-rata weight is 1, loan gets full value."""
    loans = _loans_enriched([("L1", 1, 120.0)])
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 200.0)])
    cls = _classification([(1, TYPE_1)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    assert out.height == 1
    row = out.row(0, named=True)
    assert row["aggregated_property_value"] == 200.0
    assert row["main_property_id"] == "P1"
    assert row["structure_type"] == TYPE_1


def test_flatten_type2_single_loan_multiple_properties_by_loan() -> None:
    """1 loan -> 2 properties (by_loan): loan gets sum of unique values
    (75 + 125 = 200)."""
    loans = _loans_enriched([("L1", 2, 150.0)])
    props = _properties_enriched(
        [("L1", "P_A", 2, "FIEI", 75.0), ("L1", "P_B", 2, "FIEI", 125.0)]
    )
    cls = _classification([(2, TYPE_2)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    assert out.row(0, named=True)["aggregated_property_value"] == 200.0


def test_flatten_type3_two_loans_one_property_by_loan_pro_rata() -> None:
    """2L+1P (by_loan): same property valuation duplicated per loan ->
    unique value = 400 (first non-null). Pro-rata: L3A gets 400 * 100/400
    = 100, L3B gets 400 * 300/400 = 300."""
    loans = _loans_enriched([("L3A", 3, 100.0), ("L3B", 3, 300.0)])
    props = _properties_enriched(
        [("L3A", "P3", 3, "FIEI", 400.0), ("L3B", "P3", 3, "FIEI", 400.0)]
    )
    cls = _classification([(3, TYPE_3)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    by_loan = {r["calc_loan_id"]: r["aggregated_property_value"] for r in out.iter_rows(named=True)}
    assert math.isclose(by_loan["L3A"], 100.0, abs_tol=1e-9)
    assert math.isclose(by_loan["L3B"], 300.0, abs_tol=1e-9)


def test_flatten_type4_full_set_by_loan() -> None:
    """2L+2P full set (by_loan): each property has duplicated valuation
    per loan. Unique values: P_A=200, P_B=300. Group total = 500.
    Equal CPBs (200, 200) -> each loan gets 250."""
    loans = _loans_enriched([("L4A", 4, 200.0), ("L4B", 4, 200.0)])
    props = _properties_enriched(
        [
            ("L4A", "P_A", 4, "FIEI", 200.0),
            ("L4A", "P_B", 4, "FIEI", 300.0),
            ("L4B", "P_A", 4, "FIEI", 200.0),
            ("L4B", "P_B", 4, "FIEI", 300.0),
        ]
    )
    cls = _classification([(4, TYPE_4)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    vals = sorted(r["aggregated_property_value"] for r in out.iter_rows(named=True))
    assert vals == [250.0, 250.0]


def test_flatten_type5_cross_collateralised_set_by_loan() -> None:
    """2L+2P cross set (by_loan, missing one edge):
      P_A: only L5A links -> unique value 200 (single observation).
      P_B: linked from L5A only -> unique value 100.
      Group total = 300. CPBs (100, 200): L5A gets 100, L5B gets 200."""
    loans = _loans_enriched([("L5A", 5, 100.0), ("L5B", 5, 200.0)])
    props = _properties_enriched(
        [
            ("L5A", "P_A", 5, "FIEI", 200.0),
            ("L5B", "P_A", 5, "FIEI", 200.0),
            ("L5A", "P_B", 5, "FIEI", 100.0),
            # NB: no L5B-P_B link -> "cross-collateralised" missing edge
        ]
    )
    cls = _classification([(5, TYPE_5)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    by_loan = {r["calc_loan_id"]: r["aggregated_property_value"] for r in out.iter_rows(named=True)}
    # Pro-rata float arithmetic produces ~1e-14 error; use isclose.
    assert math.isclose(by_loan["L5A"], 100.0, abs_tol=1e-9)
    assert math.isclose(by_loan["L5B"], 200.0, abs_tol=1e-9)


def test_flatten_type2_by_group_aggregation_sum() -> None:
    """1L+2P by_group: same property reported twice with SPLIT values.
    Mirrors test-flattening.R::"flatten_loan_collateral handles by_group
    aggregation correctly". Each property's split values sum: P1=100+100,
    P2=50+50. Group total = 200+100=300. Loan gets full 300."""
    loans = _loans_enriched([("L1", 1, 100.0)])
    props = _properties_enriched(
        [
            ("L1", "P1", 1, "FIEI", 100.0),
            ("L1", "P1", 1, "FIEI", 100.0),
            ("L1", "P2", 1, "FIEI", 50.0),
            ("L1", "P2", 1, "FIEI", 50.0),
        ]
    )
    cls = _classification([(1, TYPE_2)])
    out = flatten_loan_collateral(loans, props, cls, "by_group")
    assert out.row(0, named=True)["aggregated_property_value"] == 300.0


# ---------------------------------------------------------------------------
# Equal-split fallback (and NaN trigger - per user direction)
# ---------------------------------------------------------------------------


def test_flatten_equal_split_fires_when_group_total_balance_is_zero() -> None:
    """Group with all-zero CPBs -> total_loan_balance = 0 -> equal split.
    1 loan in group, total property value 100k -> loan gets 100k."""
    loans = _loans_enriched([("L1", 1, 0.0)])
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100000.0)])
    cls = _classification([(1, TYPE_1)])
    with pytest.warns(UserWarning, match="zero total balance"):
        out = flatten_loan_collateral(loans, props, cls, "by_loan")
    assert out.row(0, named=True)["aggregated_property_value"] == 100000.0


def test_flatten_equal_split_fires_when_group_total_balance_is_null() -> None:
    """Polars null in CPB -> drop_nans()+sum gives 0 over the group ->
    equal-split fallback fires. Pinned because Polars sum on all-null
    Float64 yields 0 (not null), so the trigger uses == 0 OR is_null
    OR is_nan to be defensive."""
    loans = _loans_enriched([("L1", 1, 100.0)])
    # Override CPB to null on L1
    loans = loans.with_columns(
        current_principal_balance=pl.lit(None, dtype=pl.Float64)
    )
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100000.0)])
    cls = _classification([(1, TYPE_1)])
    with pytest.warns(UserWarning, match="zero total balance"):
        out = flatten_loan_collateral(loans, props, cls, "by_loan")
    # Equal-split: 100000 / 1 loan = 100000
    assert out.row(0, named=True)["aggregated_property_value"] == 100000.0


def test_flatten_equal_split_fires_when_group_total_balance_is_nan() -> None:
    """User-direction test: Polars' is_null() does NOT catch NaN.
    R's is.na() catches BOTH. If a CPB column ever contains NaN
    (e.g. produced by upstream division by zero), the equal-split
    trigger needs `is_null() | is_nan()` to fire. Without that, the
    pro-rata branch would fire and produce NaN per-loan
    aggregated_property_value, drifting silently from R.

    Build a 2-loan group where one CPB is NaN. With drop_nans()+sum,
    total_loan_balance = the other loan's CPB (a finite number). So
    actually the trigger does NOT fire here - the pro-rata branch
    runs with the finite total. The NaN-balance loan gets NaN
    (NaN / total * group_total = NaN). The non-NaN loan gets the
    expected pro-rata share.

    For the fallback to fire on NaN, ALL loans in the group must
    have NaN CPB. That's the case this test constructs.
    """
    loans = _loans_enriched([("L1", 1, 0.0), ("L2", 1, 0.0)])
    # Override BOTH CPBs to NaN.
    loans = loans.with_columns(
        current_principal_balance=pl.Series(
            [float("nan"), float("nan")], dtype=pl.Float64
        )
    )
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100000.0)])
    cls = _classification([(1, TYPE_3)])
    with pytest.warns(UserWarning, match="zero total balance"):
        out = flatten_loan_collateral(loans, props, cls, "by_loan")
    vals = [r["aggregated_property_value"] for r in out.iter_rows(named=True)]
    # Equal split: 100000 / 2 loans = 50000 each.
    assert vals == [50000.0, 50000.0]


# ---------------------------------------------------------------------------
# Pro-rata conservation invariant (per user direction)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "fixture_name,balances,property_total",
    [
        # Pro-rata branch: 3 loans with distinct balances
        ("pro_rata_distinct", (100.0, 200.0, 300.0), 600.0),
        # Pro-rata branch: 2 loans with equal balances
        ("pro_rata_equal", (250.0, 250.0), 500.0),
        # Equal-split branch: all-zero balances
        ("equal_split_zero", (0.0, 0.0, 0.0), 900.0),
        # Equal-split branch: 1 loan
        ("trivial_one_loan", (1000.0,), 100.0),
    ],
)
def test_flatten_pro_rata_conservation_per_group(
    fixture_name: str,
    balances: tuple[float, ...],
    property_total: float,
) -> None:
    """Conservation: sum of aggregated_property_value over loans in a
    group EQUALS group_property_total. True for BOTH branches:

      Pro-rata: weights = balance_i / sum(balances), so
                sum(weights) = 1, so sum(weights * total) = total.
      Equal-split: each of n loans gets total/n, so sum = n * (total/n)
                   = total.

    This is the semantic content of pro-rata; testing it directly is
    stronger than testing specific cell values because a refactor that
    breaks the math without changing individual cells enough to fail
    synthetic-fixture parity would still violate conservation here.
    """
    n = len(balances)
    loans = _loans_enriched(
        [(f"L{i}", 1, balances[i]) for i in range(n)]
    )
    # Single property in group, valuation_amount = property_total.
    props = _properties_enriched([("L0", "P1", 1, "FIEI", property_total)])
    cls = _classification([(1, TYPE_3 if n > 1 else TYPE_1)])

    if all(b == 0.0 for b in balances):
        with pytest.warns(UserWarning):
            out = flatten_loan_collateral(loans, props, cls, "by_loan")
    else:
        out = flatten_loan_collateral(loans, props, cls, "by_loan")

    total = sum(
        r["aggregated_property_value"]
        for r in out.iter_rows(named=True)
        if not (
            isinstance(r["aggregated_property_value"], float)
            and math.isnan(r["aggregated_property_value"])
        )
    )
    assert math.isclose(total, property_total, abs_tol=1e-9), (
        f"{fixture_name}: conservation violated. "
        f"Sum of aggregated_property_value = {total}, expected {property_total}"
    )


# ---------------------------------------------------------------------------
# final_valuation_method/date/amount propagation - the B-4 anchor test
# ---------------------------------------------------------------------------


def test_flatten_propagates_final_valuation_method_date_amount_together_after_stage2_flip() -> None:
    """The MILAN cols 73/74 + Property Value contract: when Stage-2
    flips the valuation choice (current FIEI valid method but bad amount
    -> use original), ALL THREE coupled values - method, date, AND the
    amount driving aggregated_property_value - must come from the
    ORIGINAL source on the main-property record.

    This test pins the orchestrator's propagation. B-1 already verified
    select_valuation_fields' per-row output is coupled (the asymmetric
    test); B-4 verifies flatten_loan_collateral carries that through to
    the main_property_details join AND that the per-loan
    aggregated_property_value is computed from the FLIPPED original
    amount, not the bad current amount.

    Fixture:
      - 1 loan, 1 property in group 1.
      - current_valuation_method = FIEI (full inspection - Stage 1 picks current)
      - current_valuation_amount = 5.0 (<= MIN_VALID_PROPERTY_VALUE -> Stage 2 flips)
      - original_valuation_method = AUVM
      - original_valuation_date = 2019-06-15
      - original_valuation_amount = 320000.0
      - current_valuation_date = 2024-01-01 (would be wrong if not flipped)

    Expectations:
      final_valuation_method = "AUVM" (from original)
      final_valuation_date = 2019-06-15 (from original)
      aggregated_property_value = 320000.0 (from original; not 5.0)
    """
    loans = _loans_enriched([("L1", 1, 100000.0)])
    props = pl.DataFrame(
        {
            "underlying_exposure_identifier": pl.Series(["L1"], dtype=pl.String),
            "calc_property_id": pl.Series(["P1"], dtype=pl.String),
            "collateral_group_id": pl.Series([1], dtype=pl.Int64),
            "current_valuation_method": pl.Series(["FIEI"], dtype=pl.String),
            "current_valuation_amount": pl.Series([5.0], dtype=pl.Float64),
            "current_valuation_date": pl.Series([date(2024, 1, 1)], dtype=pl.Date),
            "original_valuation_method": pl.Series(["AUVM"], dtype=pl.String),
            "original_valuation_amount": pl.Series([320000.0], dtype=pl.Float64),
            "original_valuation_date": pl.Series([date(2019, 6, 15)], dtype=pl.Date),
            "occupancy_type": pl.Series(["FOWN"], dtype=pl.String),
        }
    )
    cls = _classification([(1, TYPE_1)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    row = out.row(0, named=True)

    # All three coupled values must be from the ORIGINAL source.
    assert row["final_valuation_method"] == "AUVM"
    assert row["final_valuation_date"] == date(2019, 6, 15)
    # The property total used for aggregated_property_value comes from
    # the Stage-2-flipped original amount (320000), not the bad current
    # amount (5). Single loan in group -> pro-rata weight = 1 -> agg = 320000.
    assert row["aggregated_property_value"] == 320000.0


def test_flatten_propagates_final_valuation_when_stage2_keeps_current() -> None:
    """Control: when Stage 2 does NOT flip (current valuation valid),
    the main-property record's final_* fields are from CURRENT, and
    aggregated_property_value uses the current amount. This pins the
    'no-flip' branch alongside the asymmetric flip test to prove the
    propagation actually depends on the Stage-2 selection, not on a
    hard-coded preference."""
    loans = _loans_enriched([("L1", 1, 100000.0)])
    props = pl.DataFrame(
        {
            "underlying_exposure_identifier": pl.Series(["L1"], dtype=pl.String),
            "calc_property_id": pl.Series(["P1"], dtype=pl.String),
            "collateral_group_id": pl.Series([1], dtype=pl.Int64),
            "current_valuation_method": pl.Series(["FIEI"], dtype=pl.String),
            "current_valuation_amount": pl.Series([450000.0], dtype=pl.Float64),
            "current_valuation_date": pl.Series([date(2024, 1, 1)], dtype=pl.Date),
            "original_valuation_method": pl.Series(["AUVM"], dtype=pl.String),
            "original_valuation_amount": pl.Series([320000.0], dtype=pl.Float64),
            "original_valuation_date": pl.Series([date(2019, 6, 15)], dtype=pl.Date),
            "occupancy_type": pl.Series(["FOWN"], dtype=pl.String),
        }
    )
    cls = _classification([(1, TYPE_1)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    row = out.row(0, named=True)
    assert row["final_valuation_method"] == "FIEI"
    assert row["final_valuation_date"] == date(2024, 1, 1)
    assert row["aggregated_property_value"] == 450000.0


# ---------------------------------------------------------------------------
# Hard errors
# ---------------------------------------------------------------------------


def test_flatten_invalid_aggregation_method_raises() -> None:
    """Mirrors test-flattening.R::"flatten_loan_collateral validates
    aggregation_method parameter"."""
    loans = _loans_enriched([("L1", 1, 100.0)])
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100.0)])
    cls = _classification([(1, TYPE_1)])
    with pytest.raises(ValueError, match="aggregation_method"):
        flatten_loan_collateral(loans, props, cls, "invalid")  # type: ignore[arg-type]


def test_flatten_unclassified_group_raises() -> None:
    """Mirrors test-flattening.R::"flatten_loan_collateral fails on
    unclassified groups"."""
    loans = _loans_enriched([("L1", 1, 100.0)])
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100.0)])
    cls = _classification([(1, "unclassified")])
    with pytest.raises(ValueError, match="unclassified"):
        flatten_loan_collateral(loans, props, cls, "by_loan")


# ---------------------------------------------------------------------------
# Output shape
# ---------------------------------------------------------------------------


def test_flatten_output_one_row_per_loan() -> None:
    """Sanity: output has exactly one row per calc_loan_id, even when
    main_property_details join could in principle multiply rows."""
    loans = _loans_enriched([("L1", 1, 100.0), ("L2", 1, 100.0), ("L3", 2, 200.0)])
    props = _properties_enriched(
        [
            ("L1", "P1", 1, "FIEI", 100.0),
            ("L2", "P1", 1, "FIEI", 100.0),
            ("L3", "P2", 2, "FIEI", 200.0),
        ]
    )
    cls = _classification([(1, TYPE_3), (2, TYPE_1)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    assert out.height == 3
    assert sorted(out["calc_loan_id"].to_list()) == ["L1", "L2", "L3"]


def test_flatten_output_carries_expected_columns() -> None:
    """The flattened output carries (loans cols) + aggregated_property_value
    + main_property_details cols + structure_type."""
    loans = _loans_enriched([("L1", 1, 100.0)])
    props = _properties_enriched([("L1", "P1", 1, "FIEI", 100.0)])
    cls = _classification([(1, TYPE_1)])
    out = flatten_loan_collateral(loans, props, cls, "by_loan")
    expected_subset = {
        "calc_loan_id",
        "collateral_group_id",
        "current_principal_balance",
        "aggregated_property_value",
        "main_property_id",
        "final_valuation_method",
        "final_valuation_date",
        "structure_type",
    }
    assert expected_subset.issubset(set(out.columns))


# ---------------------------------------------------------------------------
# B-5 / Stage 8: apply_derived_fields - LTVs + seasoning + calc_ rename + reorder
# ---------------------------------------------------------------------------


def _flat_minimal(
    *,
    group_ids: Sequence[int],
    loan_ids: Sequence[str],
    current_balances: Sequence[float | None],
    original_balances: Sequence[float | None],
    aggregated_property_values: Sequence[float | None],
    pool_cutoff_dates: Sequence[date | None] | None = None,
    origination_dates: Sequence[date | None] | None = None,
) -> pl.DataFrame:
    """Minimal flatten-shaped frame for testing apply_derived_fields.

    Carries the columns apply_derived_fields touches; doesn't try to
    reproduce the full 82-column flatten output.
    """
    n = len(loan_ids)
    if pool_cutoff_dates is None:
        pool_cutoff_dates = [date(2024, 6, 30)] * n
    if origination_dates is None:
        origination_dates = [date(2020, 1, 1)] * n
    return pl.DataFrame(
        {
            "calc_loan_id": pl.Series(list(loan_ids), dtype=pl.String),
            "collateral_group_id": pl.Series(list(group_ids), dtype=pl.Int64),
            "current_principal_balance": pl.Series(
                list(current_balances), dtype=pl.Float64
            ),
            "original_principal_balance": pl.Series(
                list(original_balances), dtype=pl.Float64
            ),
            "aggregated_property_value": pl.Series(
                list(aggregated_property_values), dtype=pl.Float64
            ),
            "pool_cutoff_date": pl.Series(
                list(pool_cutoff_dates), dtype=pl.Date
            ),
            "origination_date": pl.Series(
                list(origination_dates), dtype=pl.Date
            ),
        }
    )


# -- LTV: standard + boundary cases -----------------------------------------


def test_apply_derived_fields_ltv_single_loan_group() -> None:
    """Single loan in group: LTV = balance / property value."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[200000.0],
        aggregated_property_values=[200000.0],
    )
    out = apply_derived_fields(df)
    row = out.row(0, named=True)
    assert row["calc_current_LTV"] == 0.5
    assert row["calc_original_LTV"] == 1.0


def test_apply_derived_fields_ltv_multi_loan_group_shared_denominator() -> None:
    """Multi-loan group: BOTH LTVs share the same denominator
    (sum of aggregated_property_value across the group)."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 1],
        loan_ids=["L1", "L2"],
        current_balances=[100000.0, 80000.0],
        original_balances=[120000.0, 100000.0],
        aggregated_property_values=[200000.0, 200000.0],
    )
    out = apply_derived_fields(df)
    rows = out.to_dicts()
    # Denom = 200k + 200k = 400k. CPB sum = 180k. OPB sum = 220k.
    for r in rows:
        assert r["calc_current_LTV"] == 0.45
        assert r["calc_original_LTV"] == 0.55


def test_apply_derived_fields_ltv_nan_in_numerator_coalesces_to_zero() -> None:
    """User-direction asymmetry test: NaN in current_principal_balance
    (numerator term) coalesces to 0 inside the sum but does NOT
    collapse the LTV. Other rows in the group still contribute."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 1],
        loan_ids=["L1", "L2"],
        current_balances=[float("nan"), 100000.0],  # one NaN numerator
        original_balances=[120000.0, 100000.0],
        aggregated_property_values=[200000.0, 200000.0],
    )
    out = apply_derived_fields(df)
    # NaN row contributes 0 to current sum -> CPB sum = 0 + 100k = 100k.
    # Denom = 400k. LTV = 0.25.
    rows = out.to_dicts()
    for r in rows:
        assert r["calc_current_LTV"] == 0.25
        assert r["calc_original_LTV"] == 0.55


def test_apply_derived_fields_ltv_one_nan_denominator_does_not_collapse() -> None:
    """One row with NaN aggregated_property_value: that row contributes
    0 to the denom sum, but other rows in the group still contribute.
    LTV is still computable."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 1],
        loan_ids=["L1", "L2"],
        current_balances=[100000.0, 80000.0],
        original_balances=[100000.0, 80000.0],
        aggregated_property_values=[float("nan"), 200000.0],  # one NaN denom
    )
    out = apply_derived_fields(df)
    rows = out.to_dicts()
    # Denom sum = 0 + 200k = 200k. CPB sum = 180k. LTV = 0.9.
    for r in rows:
        assert r["calc_current_LTV"] == 0.9


def test_apply_derived_fields_ltv_all_nan_denominator_collapses_to_null() -> None:
    """User-direction third-addition boundary: ALL rows in the group
    have NaN aggregated_property_value. After fill_nan(0).sum() the
    denominator is 0, the > 0 gate fails, LTV collapses to null.
    Different code path from explicit-zero-denom but same outcome.
    """
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 1],
        loan_ids=["L1", "L2"],
        current_balances=[100000.0, 80000.0],
        original_balances=[100000.0, 80000.0],
        aggregated_property_values=[float("nan"), float("nan")],
    )
    out = apply_derived_fields(df)
    rows = out.to_dicts()
    for r in rows:
        assert r["calc_current_LTV"] is None
        assert r["calc_original_LTV"] is None


def test_apply_derived_fields_ltv_zero_denominator_collapses_to_null() -> None:
    """Explicit zero denominator: denom-gate fails -> LTV is null."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[100000.0],
        aggregated_property_values=[0.0],
    )
    out = apply_derived_fields(df)
    row = out.row(0, named=True)
    assert row["calc_current_LTV"] is None
    assert row["calc_original_LTV"] is None


def test_apply_derived_fields_ltv_null_in_numerator_coalesces_to_zero() -> None:
    """Polars null (vs NaN) in current_principal_balance: same behaviour
    as NaN under our fill_null(0).fill_nan(0) chain."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 1],
        loan_ids=["L1", "L2"],
        current_balances=[None, 100000.0],
        original_balances=[120000.0, 100000.0],
        aggregated_property_values=[200000.0, 200000.0],
    )
    out = apply_derived_fields(df)
    row = out.row(0, named=True)
    assert row["calc_current_LTV"] == 0.25


# -- Seasoning --------------------------------------------------------------


def test_apply_derived_fields_seasoning_standard() -> None:
    """Standard case: (cutoff - origination) days / 365.25.

    Anchor: synthetic fixture's ORIG_001 has origination_date 2019-03-15
    and pool_cutoff_date 2024-06-30 -> 1934 days -> 5.29500342231348."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[100000.0],
        aggregated_property_values=[200000.0],
        pool_cutoff_dates=[date(2024, 6, 30)],
        origination_dates=[date(2019, 3, 15)],
    )
    out = apply_derived_fields(df)
    assert out.row(0, named=True)["calc_seasoning"] == 1934 / 365.25


def test_apply_derived_fields_seasoning_same_day_origination_is_zero() -> None:
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[100000.0],
        aggregated_property_values=[200000.0],
        pool_cutoff_dates=[date(2024, 6, 30)],
        origination_dates=[date(2024, 6, 30)],
    )
    assert apply_derived_fields(df).row(0, named=True)["calc_seasoning"] == 0.0


def test_apply_derived_fields_seasoning_future_origination_is_negative() -> None:
    """Origination AFTER pool_cutoff (analyst-input data quality issue):
    seasoning is negative, no warning, no error.

    Matches R behaviour - R doesn't warn on future-dated origination
    either. Treated as an analyst-input data quality issue rather than
    a pipeline error; the issue is logged in the R-repo tracker for
    upstream review, not blocking. The absence of a warning here is
    INTENTIONAL, not an oversight - if a future contributor adds a
    validation here, this test should fail loudly to prompt them to
    re-read the parity contract before changing behaviour.
    """
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[100000.0],
        aggregated_property_values=[200000.0],
        pool_cutoff_dates=[date(2024, 1, 1)],
        origination_dates=[date(2024, 7, 1)],  # 6 months in the future
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error", UserWarning)  # error if any warns
        out = apply_derived_fields(df)
    seasoning = out.row(0, named=True)["calc_seasoning"]
    assert seasoning is not None
    assert seasoning < 0


def test_apply_derived_fields_seasoning_null_pool_cutoff_propagates_to_null() -> None:
    """Per-row null pool_cutoff_date -> null calc_seasoning for that row.

    Verified against r_reference/R/pipeline.R:540-546: R's
    safe_as_date(NA, ...) returns NA_Date_, difftime(NA_Date_, x)
    returns NA, divided by 365.25 stays NA. No warning, no error -
    R relies on Stage 1's validate_required_columns to catch the
    column-missing case (which IS a hard error)."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1, 2],
        loan_ids=["L1", "L2"],
        current_balances=[100000.0, 100000.0],
        original_balances=[100000.0, 100000.0],
        aggregated_property_values=[200000.0, 200000.0],
        pool_cutoff_dates=[None, date(2024, 6, 30)],  # row 0 null
        origination_dates=[date(2020, 1, 1), date(2020, 1, 1)],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error", UserWarning)
        out = apply_derived_fields(df)
    rows = out.to_dicts()
    assert rows[0]["calc_seasoning"] is None
    assert rows[1]["calc_seasoning"] is not None
    # The non-null row computes normally.
    assert rows[1]["calc_seasoning"] == (date(2024, 6, 30) - date(2020, 1, 1)).days / 365.25


def test_apply_derived_fields_seasoning_null_origination_propagates_to_null() -> None:
    """Per-row null origination_date: same null-propagation as null
    pool_cutoff_date - difftime(x, NA) = NA."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    df = _flat_minimal(
        group_ids=[1],
        loan_ids=["L1"],
        current_balances=[100000.0],
        original_balances=[100000.0],
        aggregated_property_values=[200000.0],
        pool_cutoff_dates=[date(2024, 6, 30)],
        origination_dates=[None],
    )
    out = apply_derived_fields(df)
    assert out.row(0, named=True)["calc_seasoning"] is None


# -- calc_ rename -----------------------------------------------------------


def test_apply_derived_fields_renames_to_calc_prefix() -> None:
    """The 6 rename-with-calc-prefix columns + 2 custom-rename columns
    should all flip to their calc_ names. R: pipeline.R:553-559."""
    from esma_milan.pipeline.flatten import apply_derived_fields

    # Build a frame with the columns that get renamed. Add the LTV/seasoning
    # input columns too so the helper doesn't error.
    df = pl.DataFrame(
        {
            "calc_loan_id": ["L1"],
            "current_principal_balance": [100.0],
            "original_principal_balance": [100.0],
            "pool_cutoff_date": pl.Series([date(2024, 1, 1)], dtype=pl.Date),
            "origination_date": pl.Series([date(2020, 1, 1)], dtype=pl.Date),
            # Rename targets:
            "main_property_id": ["P1"],
            "structure_type": ["1: one loan → one property"],
            "collateral_group_id": pl.Series([1], dtype=pl.Int64),
            "cross_collateralized_set": [False],
            "full_set": [False],
            "aggregated_property_value": [200.0],
            "loans": pl.Series([1], dtype=pl.Int64),
            "collaterals": pl.Series([1], dtype=pl.Int64),
        }
    )
    out = apply_derived_fields(df)
    # Pre-rename names must NOT survive.
    for old in (
        "main_property_id",
        "structure_type",
        "collateral_group_id",
        "cross_collateralized_set",
        "full_set",
        "aggregated_property_value",
        "loans",
        "collaterals",
    ):
        assert old not in out.columns, f"unrenamed: {old}"
    # Post-rename names must all be present.
    for new in (
        "calc_main_property_id",
        "calc_structure_type",
        "calc_collateral_group_id",
        "calc_cross_collateralized_set",
        "calc_full_set",
        "calc_aggregated_property_value",
        "calc_nr_loans_in_group",
        "calc_nr_properties_in_group",
    ):
        assert new in out.columns, f"missing after rename: {new}"


# -- Column reorder: exact match against the 85-col Sheet 9 fixture ---------


def test_apply_derived_fields_full_pipeline_matches_fixture_column_order() -> None:
    """The 85-column Sheet 9 layout is the parity contract. Pinned via
    the full pipeline against the synthetic fixture: any drift in
    apply_derived_fields' rename/reorder logic surfaces here as
    "column N: actual=X expected=Y" rather than buried in a cell-by-cell
    parity diff failure."""
    import openpyxl

    from esma_milan.pipeline.classification import run_stage5
    from esma_milan.pipeline.enriched import (
        compose_loans_enriched,
        compose_properties_enriched,
    )
    from esma_milan.pipeline.filters import run_stage2
    from esma_milan.pipeline.flatten import run_stage7
    from esma_milan.pipeline.graph import run_stage4
    from esma_milan.pipeline.identifiers import run_stage3
    from esma_milan.pipeline.stage1 import run_stage1

    repo_root = Path(__file__).resolve().parents[2]
    syn = repo_root / "tests" / "fixtures" / "synthetic"

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        s1 = run_stage1(
            loans_path=syn / "loans.csv",
            collaterals_path=syn / "collaterals.csv",
            taxonomy_path=syn / "taxonomy.xlsx",
        )
        s2 = run_stage2(s1.loans, s1.properties)
        s3 = run_stage3(s2.loans, s2.properties)
        s4 = run_stage4(s3.loans, s3.properties)
        s5 = run_stage5(s4)
        le = compose_loans_enriched(s3.loans, s4.loan_groups, s5.classifications)
        pe = compose_properties_enriched(
            s3.properties, s4.collateral_groups, s5.classifications
        )
        s7 = run_stage7(le, pe, s5.classifications, aggregation_method="by_loan")

    py_cols = list(s7.combined_flattened.columns)

    wb = openpyxl.load_workbook(
        syn / "expected_r_output.xlsx", read_only=True, data_only=True
    )
    ws = wb["Combined flattened pool"]
    ws.reset_dimensions()
    fixture_cols = list(next(ws.iter_rows(values_only=True)))

    assert len(py_cols) == 85, f"expected 85 cols, got {len(py_cols)}"
    assert py_cols == fixture_cols, (
        "column order drift vs Sheet 9 fixture:\n"
        + "\n".join(
            f"  [{i:2d}] py={p!r:50s} fix={f!r}"
            for i, (p, f) in enumerate(zip(py_cols, fixture_cols, strict=True))
            if p != f
        )
    )


