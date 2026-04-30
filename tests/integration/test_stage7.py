"""Stage 7 / B-5 integration test: end-to-end "Combined flattened pool" output.

Pins the Sheet 9 contract against the staged synthetic fixture:
- 85-col output with byte-equal column order
- LTV / seasoning values for ORIG_001 anchored against the R fixture
- Group-level LTV consistency (all loans in a group share the same value)
- final_valuation_method / final_valuation_date propagation through the
  orchestrator (asymmetric Stage-1/Stage-2 flip case verified by B-4 unit
  tests; this integration check confirms the synthetic pool's main
  properties carry through correctly)
"""

from __future__ import annotations

import warnings
from pathlib import Path

import openpyxl
import pytest

from esma_milan.runner import PipelineResult, run_pipeline

REPO_ROOT = Path(__file__).resolve().parents[2]
SYNTHETIC = REPO_ROOT / "tests" / "fixtures" / "synthetic"


@pytest.fixture(scope="module")
def synthetic_pipeline_run(tmp_path_factory: pytest.TempPathFactory) -> PipelineResult:
    out_dir = tmp_path_factory.mktemp("stage7_synthetic")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return run_pipeline(
            loans_file_path=SYNTHETIC / "loans.csv",
            collaterals_file_path=SYNTHETIC / "collaterals.csv",
            taxonomy_file_path=SYNTHETIC / "taxonomy.xlsx",
            deal_name="SYNTHETIC_FIXTURE",
            output_dir=out_dir,
            verbose=False,
        )


def _fixture_sheet9() -> tuple[list[str], list[tuple[object, ...]]]:
    wb = openpyxl.load_workbook(
        SYNTHETIC / "expected_r_output.xlsx", read_only=True, data_only=True
    )
    ws = wb["Combined flattened pool"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[0]), rows[1:]


def test_stage7_combined_flattened_present(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    assert synthetic_pipeline_run.stage7 is not None
    assert synthetic_pipeline_run.stage7.combined_flattened.height == 8
    assert synthetic_pipeline_run.stage7.combined_flattened.width == 85


def test_stage7_column_order_matches_fixture(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    """End-to-end column order pinning - mirrors the unit-level test in
    test_flatten.py but runs through the full PipelineResult to catch
    any drift introduced by runner-side composition."""
    assert synthetic_pipeline_run.stage7 is not None
    fixture_cols, _ = _fixture_sheet9()
    assert list(synthetic_pipeline_run.stage7.combined_flattened.columns) == fixture_cols


def test_stage7_orig_001_derived_values_match_fixture(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    """Anchor on ORIG_001 (NEW_001):
      - calc_current_LTV  = 0.5625    (180000 / 320000)
      - calc_original_LTV = 0.6875    (220000 / 320000)
      - calc_seasoning    = 5.29500342231348  ((2024-06-30 - 2019-03-15)/365.25)
    """
    assert synthetic_pipeline_run.stage7 is not None
    df = synthetic_pipeline_run.stage7.combined_flattened
    row = df.filter(df["calc_loan_id"] == "NEW_001").row(0, named=True)
    assert row["calc_current_LTV"] == pytest.approx(0.5625, abs=1e-12)
    assert row["calc_original_LTV"] == pytest.approx(0.6875, abs=1e-12)
    assert row["calc_seasoning"] == pytest.approx(5.29500342231348, abs=1e-12)


def test_stage7_group_level_ltv_consistency(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    """All loans in the same collateral_group_id share the same
    calc_current_LTV and calc_original_LTV (group-level computation)."""
    import polars as pl

    assert synthetic_pipeline_run.stage7 is not None
    df = synthetic_pipeline_run.stage7.combined_flattened
    grouped = df.group_by("calc_collateral_group_id").agg(
        n_unique_curr=pl.col("calc_current_LTV").n_unique(),
        n_unique_orig=pl.col("calc_original_LTV").n_unique(),
    )
    # Each group has 1 distinct LTV value (or 1 distinct null).
    for n in grouped["n_unique_curr"].to_list():
        assert n == 1, "calc_current_LTV varies within a group"
    for n in grouped["n_unique_orig"].to_list():
        assert n == 1, "calc_original_LTV varies within a group"


def test_stage7_final_valuation_propagation_to_synthetic_main_properties(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    """The final_valuation_method/date columns MUST match the values from
    the chosen Stage-2 final pick on each main property.

    Synthetic-fixture spot-check: NEW_002 belongs to group 2 whose main
    property (per the analyst-walkthrough) is PORIG_02 with FIEI current
    method and 2024-01-15 valuation date... let's pull from the fixture
    rather than hard-coding."""
    assert synthetic_pipeline_run.stage7 is not None
    df = synthetic_pipeline_run.stage7.combined_flattened
    fixture_cols, fixture_rows = _fixture_sheet9()
    fix_idx = {c: i for i, c in enumerate(fixture_cols)}

    py_by_loan = {
        row["calc_loan_id"]: row for row in df.iter_rows(named=True)
    }
    fix_by_loan = {
        r[fix_idx["calc_loan_id"]]: r for r in fixture_rows
    }

    for loan_id in py_by_loan:
        py = py_by_loan[loan_id]
        fix = fix_by_loan[loan_id]
        # Compare each of the propagation columns.
        assert py["final_valuation_method"] == fix[fix_idx["final_valuation_method"]], (
            f"{loan_id}: final_valuation_method drift"
        )
        # final_valuation_date - py is a Date object; fixture is an Excel
        # serial int. Convert via the same epoch constant the writer uses.
        from esma_milan.io_layer.write_workbook import _date_to_excel_serial
        py_date = py["final_valuation_date"]
        fix_serial = fix[fix_idx["final_valuation_date"]]
        if py_date is None and fix_serial is None:
            continue
        assert _date_to_excel_serial(py_date) == fix_serial, (
            f"{loan_id}: final_valuation_date drift"
        )


def test_stage7_workbook_writes_combined_flattened_pool_sheet(
    synthetic_pipeline_run: PipelineResult,
) -> None:
    """Sanity check that the workbook actually contains Sheet 9 with
    85 cols + 8 data rows."""
    assert synthetic_pipeline_run.output_path is not None
    wb = openpyxl.load_workbook(
        synthetic_pipeline_run.output_path, read_only=True, data_only=True
    )
    ws = wb["Combined flattened pool"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 9, "expected 1 header + 8 data rows"
    assert len(rows[0]) == 85, "expected 85 columns"
