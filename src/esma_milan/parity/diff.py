"""Cell-by-cell workbook diff.

Compares two .xlsx files produced by the ESMA -> MILAN pipeline. Tolerances
mirror §6.2 of the project brief:
  * Numeric: 1e-9 absolute, 1e-12 relative.
  * String: byte-equal (no normalisation other than openpyxl's own).
  * Date: ISO-string-equal in MILAN sheet, Excel-serial-equal elsewhere.
    The pipeline writer determines the storage format per sheet via
    config.SHEETS_WITH_EXCEL_SERIAL_DATES; the diff just compares cells
    in whatever form they come out of openpyxl, which preserves both.

Group-id canonicalisation has two layers:

  1. In-sheet relabel: for sheets that carry both a row-unique
     identifier (calc_loan_id / Loan Identifier / calc_property_id /
     ...) AND a collateral_group_id column, the diff builds a per-side
     {gid -> 1..N rank} map from those rows and relabels gid CELL
     VALUES on cell comparison so R and Python's different integer
     labels for the same partition compare equal.

  2. Cross-sheet WorkbookContext (this stage's addition): for sheets
     whose own rows do not carry a row-unique identifier (notably
     "Group classifications", which is one row per group with no
     loan/property column), the diff builds a context from the
     "Cleaned ESMA loans" sheet's (calc_loan_id, collateral_group_id)
     columns and uses it to:
       (a) relabel gid CELL VALUES on cell comparison, and
       (b) reorder ROWS on each side by the lex-smallest calc_loan_id
           in the group, so the same partition produces matching row
           order regardless of which integer labels each side picked.

  The cross-sheet alignment only fires when BOTH workbooks' contexts
  are populated. If either side has an empty Cleaned ESMA loans sheet
  (e.g. a partial pipeline output where Stage 5 has landed but the
  Cleaned ESMA loans writer hasn't), the alignment is a no-op and the
  diff falls back to positional row matching. This keeps behaviour
  predictable while later stages stack up; the unit tests in
  tests/unit/test_diff.py exercise both branches with constructed
  fixtures.
"""

from __future__ import annotations

import math
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

NUMERIC_ABS_TOL: float = 1e-9
NUMERIC_REL_TOL: float = 1e-12

# Columns whose values are arbitrary integer group labels and need
# canonicalising before comparison. The values themselves are meaningless;
# only the partition they induce matters.
GROUP_ID_COLUMNS: frozenset[str] = frozenset({
    "collateral_group_id",
    "calc_collateral_group_id",
    "Additional data 4 - calc_collateral_group_id",
})

# When canonicalising group IDs we need to know which column holds a
# row-unique identifier per row, so we can pick "sorted min identifier"
# as the canonical group label. Different sheets use different
# row-identifier columns; the priority list is searched in order and
# the first present column wins.
NODE_ID_COLUMNS_BY_PRIORITY: tuple[str, ...] = (
    # Loan-side row identifiers (Cleaned ESMA loans, Combined flattened
    # pool, MILAN template pool).
    "calc_loan_id",
    "Loan Identifier",
    "Additional data 1 - calc_loan_id",
    "loan_id",
    "new_underlying_exposure_identifier",
    # Property-side row identifiers (Cleaned ESMA properties). The
    # property sheet has no calc_loan_id column, so we fall through to
    # property-id columns. underlying_exposure_identifier is NOT
    # row-unique on the property sheet (a single loan can back multiple
    # properties), so calc_property_id is preferred.
    "calc_property_id",
    "Property Identifier",
    "Additional data 3 - calc_main_property_id",
)

# Sheets whose rows do NOT carry a row-unique row identifier of their own
# and therefore need cross-sheet alignment via the WorkbookContext built
# from the Cleaned ESMA loans sheet. Currently just "Group classifications"
# (one row per group).
SHEETS_NEEDING_CROSS_SHEET_ALIGNMENT: frozenset[str] = frozenset({
    "Group classifications",
})

# The sheet WorkbookContext is derived from. Single-source for now;
# extending this to a priority list (e.g. fall back to "Combined
# flattened pool" if Cleaned ESMA loans is empty) is straightforward
# but unneeded until the writer for those sheets ships.
LOANS_CONTEXT_SHEET: str = "Cleaned ESMA loans"
LOANS_CONTEXT_LOAN_ID_COLUMN: str = "calc_loan_id"
LOANS_CONTEXT_GID_COLUMN: str = "collateral_group_id"


@dataclass(frozen=True)
class CellDiff:
    """A single mismatched cell."""

    sheet: str
    row: int  # 1-indexed; row 1 is the header row, data rows start at 2.
    column_index: int  # 1-indexed within the sheet.
    column_name: str
    actual: Any
    expected: Any
    reason: str

    def short(self) -> str:
        a = _short_repr(self.actual)
        e = _short_repr(self.expected)
        return (
            f"  [{self.sheet}!{_col_letter(self.column_index)}{self.row} "
            f"{self.column_name!r}] {self.reason}: actual={a} expected={e}"
        )


@dataclass
class SheetDiff:
    """All differences in a single sheet."""

    sheet: str
    actual_present: bool = True
    expected_present: bool = True
    actual_dimensions: tuple[int, int] = (0, 0)  # (rows, cols)
    expected_dimensions: tuple[int, int] = (0, 0)
    column_order_match: bool = True
    column_order_actual: tuple[str, ...] = ()
    column_order_expected: tuple[str, ...] = ()
    cell_diffs: list[CellDiff] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return (
            self.actual_present
            and self.expected_present
            and self.actual_dimensions == self.expected_dimensions
            and self.column_order_match
            and not self.cell_diffs
        )


@dataclass
class DiffReport:
    """Aggregate result of comparing two workbooks."""

    actual_path: Path
    expected_path: Path
    sheet_diffs: list[SheetDiff] = field(default_factory=list)
    actual_sheet_order: tuple[str, ...] = ()
    expected_sheet_order: tuple[str, ...] = ()

    @property
    def passed(self) -> bool:
        return (
            self.actual_sheet_order == self.expected_sheet_order
            and all(s.passed for s in self.sheet_diffs)
        )

    def summary(self) -> dict[str, Any]:
        return {
            "passed": self.passed,
            "sheets": [
                {
                    "name": s.sheet,
                    "passed": s.passed,
                    "actual_dim": s.actual_dimensions,
                    "expected_dim": s.expected_dimensions,
                    "column_order_match": s.column_order_match,
                    "cell_diffs": len(s.cell_diffs),
                }
                for s in self.sheet_diffs
            ],
        }


@dataclass(frozen=True)
class WorkbookContext:
    """Cross-sheet info derived from one workbook's Cleaned ESMA loans sheet.

    Empty if the loans sheet is missing, has no rows, or is missing the
    expected (calc_loan_id, collateral_group_id) columns. The diff falls
    back to in-sheet relabel + positional row matching when either side's
    context is empty - see SHEETS_NEEDING_CROSS_SHEET_ALIGNMENT handling
    in `_diff_sheet`.
    """

    gid_to_canonical: dict[Any, int]
    """gid (this workbook's label space) -> 1..N canonical rank, where
    rank is determined by lex-smallest calc_loan_id in the group."""

    gid_to_min_loan_id: dict[Any, str]
    """gid (this workbook's label space) -> lex-smallest calc_loan_id in
    that group. Used as the row-sort key for SHEETS_NEEDING_CROSS_SHEET_
    ALIGNMENT."""

    @property
    def is_populated(self) -> bool:
        return bool(self.gid_to_min_loan_id)


_EMPTY_CONTEXT: WorkbookContext = WorkbookContext(
    gid_to_canonical={}, gid_to_min_loan_id={}
)


def diff_workbooks(
    actual_path: Path | str,
    expected_path: Path | str,
    *,
    max_cell_diffs_per_sheet: int = 50,
) -> DiffReport:
    """Diff two .xlsx files. Always returns a report; never raises on
    diff. Raises only on file-level errors (file missing, malformed)."""
    actual_path = Path(actual_path)
    expected_path = Path(expected_path)

    actual_wb = _load(actual_path)
    expected_wb = _load(expected_path)

    actual_ctx = _build_workbook_context(actual_wb)
    expected_ctx = _build_workbook_context(expected_wb)

    report = DiffReport(
        actual_path=actual_path,
        expected_path=expected_path,
        actual_sheet_order=tuple(actual_wb.sheetnames),
        expected_sheet_order=tuple(expected_wb.sheetnames),
    )

    all_sheets = list(expected_wb.sheetnames)
    # Also report any extra sheets in actual that aren't in expected.
    for name in actual_wb.sheetnames:
        if name not in all_sheets:
            all_sheets.append(name)

    for sheet_name in all_sheets:
        sd = _diff_sheet(
            actual_wb,
            expected_wb,
            sheet_name,
            actual_ctx=actual_ctx,
            expected_ctx=expected_ctx,
            max_cell_diffs=max_cell_diffs_per_sheet,
        )
        report.sheet_diffs.append(sd)

    return report


def _load(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Many openxlsx-produced sheets advertise dim 1x1; force a re-scan.
    for ws in wb.worksheets:
        ws.reset_dimensions()
    return wb


def _read_sheet(ws: Worksheet) -> tuple[tuple[str, ...], list[tuple[Any, ...]]]:
    """Return (header tuple, list of data row tuples). Empty sheet -> ((), [])."""
    rows_iter = ws.iter_rows(values_only=True)
    header: tuple[Any, ...] = ()
    try:
        header = next(rows_iter)
    except StopIteration:
        return (), []
    if header is None:
        return (), []
    header_str = tuple("" if h is None else str(h) for h in header)
    data_rows = list(rows_iter)
    # openpyxl sometimes pads with trailing all-None rows; strip them.
    while data_rows and all(c is None for c in data_rows[-1]):
        data_rows.pop()
    return header_str, data_rows


def _diff_sheet(
    actual_wb: openpyxl.Workbook,
    expected_wb: openpyxl.Workbook,
    sheet_name: str,
    *,
    actual_ctx: WorkbookContext,
    expected_ctx: WorkbookContext,
    max_cell_diffs: int,
) -> SheetDiff:
    sd = SheetDiff(sheet=sheet_name)

    if sheet_name not in actual_wb.sheetnames:
        sd.actual_present = False
        sd.notes.append("sheet missing in actual")
    if sheet_name not in expected_wb.sheetnames:
        sd.expected_present = False
        sd.notes.append("sheet missing in expected")
    if not (sd.actual_present and sd.expected_present):
        return sd

    a_hdr, a_rows = _read_sheet(actual_wb[sheet_name])
    e_hdr, e_rows = _read_sheet(expected_wb[sheet_name])

    sd.actual_dimensions = (len(a_rows), len(a_hdr))
    sd.expected_dimensions = (len(e_rows), len(e_hdr))
    sd.column_order_actual = a_hdr
    sd.column_order_expected = e_hdr
    sd.column_order_match = a_hdr == e_hdr

    # If column order doesn't match, don't try to diff cells positionally;
    # the report will say "column order mismatch" loudly enough.
    if not sd.column_order_match:
        return sd

    # Cross-sheet alignment for sheets whose rows lack a row-unique row
    # identifier of their own (e.g. Group classifications). Fires only
    # when BOTH sides have a populated WorkbookContext - falls back to
    # positional row matching otherwise so partial pipeline outputs
    # (Sheet 8 written but Sheet 6 not yet) keep working.
    if (
        sheet_name in SHEETS_NEEDING_CROSS_SHEET_ALIGNMENT
        and actual_ctx.is_populated
        and expected_ctx.is_populated
    ):
        a_rows = _sort_rows_by_loan_canonical_key(a_rows, a_hdr, actual_ctx)
        e_rows = _sort_rows_by_loan_canonical_key(e_rows, e_hdr, expected_ctx)
        a_relabel: dict[Any, Any] = dict(actual_ctx.gid_to_canonical)
        e_relabel: dict[Any, Any] = dict(expected_ctx.gid_to_canonical)
        sd.notes.append(
            "rows aligned by lex-smallest calc_loan_id from Cleaned ESMA loans"
        )
    else:
        # Default: build a per-sheet relabel from in-sheet row data.
        a_relabel = _build_group_relabel(a_hdr, a_rows)
        e_relabel = _build_group_relabel(e_hdr, e_rows)

    # If row count differs, diff what we can and note the mismatch.
    if sd.actual_dimensions[0] != sd.expected_dimensions[0]:
        sd.notes.append(
            f"row-count mismatch: actual={sd.actual_dimensions[0]} "
            f"expected={sd.expected_dimensions[0]}"
        )

    n_compared_rows = min(len(a_rows), len(e_rows))
    for r_idx in range(n_compared_rows):
        a_row = a_rows[r_idx]
        e_row = e_rows[r_idx]
        for c_idx, col_name in enumerate(a_hdr):
            a_val = a_row[c_idx] if c_idx < len(a_row) else None
            e_val = e_row[c_idx] if c_idx < len(e_row) else None
            if col_name in GROUP_ID_COLUMNS:
                a_val = a_relabel.get(a_val, a_val)
                e_val = e_relabel.get(e_val, e_val)
            ok, reason = _cells_equal(a_val, e_val)
            if not ok:
                if len(sd.cell_diffs) >= max_cell_diffs:
                    sd.notes.append(
                        f"truncated at {max_cell_diffs} cell diffs for this sheet"
                    )
                    return sd
                sd.cell_diffs.append(
                    CellDiff(
                        sheet=sheet_name,
                        row=r_idx + 2,  # +2: 1-indexed and skip header row
                        column_index=c_idx + 1,
                        column_name=col_name,
                        actual=a_val,
                        expected=e_val,
                        reason=reason,
                    )
                )
    return sd


def _cells_equal(actual: Any, expected: Any) -> tuple[bool, str]:
    """Return (equal, reason). reason is empty when equal."""
    # None / "" equivalence: openxlsx writes empty cells as None; some
    # readers surface the empty string. Treat them as equal.
    a_empty = actual is None or actual == ""
    e_empty = expected is None or expected == ""
    if a_empty and e_empty:
        return True, ""
    if a_empty or e_empty:
        return False, "one side is empty"

    # Numeric tolerance comparison.
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        if isinstance(actual, bool) or isinstance(expected, bool):
            # Booleans compared as exact values, not numerics.
            return (actual == expected), ("bool mismatch" if actual != expected else "")
        if math.isnan(actual) and math.isnan(expected):
            return True, ""
        if math.isnan(actual) or math.isnan(expected):
            return False, "NaN mismatch"
        if math.isclose(
            actual, expected, abs_tol=NUMERIC_ABS_TOL, rel_tol=NUMERIC_REL_TOL
        ):
            return True, ""
        return False, f"numeric diff |{actual - expected:g}|"

    # Mixed numeric / string — this happens when one side stored a date as
    # serial number and the other as ISO string. The pipeline contract is
    # to write each sheet in its R-native format; mixed-type cells are a
    # genuine diff.
    if isinstance(actual, (int, float)) ^ isinstance(expected, (int, float)):
        return False, "type mismatch (numeric vs non-numeric)"

    # String / boolean / datetime — exact equality.
    if actual == expected:
        return True, ""
    return False, "value mismatch"


def _build_workbook_context(wb: openpyxl.Workbook) -> WorkbookContext:
    """Read the Cleaned ESMA loans sheet and derive the cross-sheet
    canonicalisation context.

    Returns _EMPTY_CONTEXT if the loans sheet is missing, has no rows,
    or doesn't carry the expected (calc_loan_id, collateral_group_id)
    columns. Callers must check `WorkbookContext.is_populated` before
    using the maps.
    """
    if LOANS_CONTEXT_SHEET not in wb.sheetnames:
        return _EMPTY_CONTEXT

    header, rows = _read_sheet(wb[LOANS_CONTEXT_SHEET])
    if not header or not rows:
        return _EMPTY_CONTEXT
    if (
        LOANS_CONTEXT_LOAN_ID_COLUMN not in header
        or LOANS_CONTEXT_GID_COLUMN not in header
    ):
        return _EMPTY_CONTEXT

    loan_idx = header.index(LOANS_CONTEXT_LOAN_ID_COLUMN)
    gid_idx = header.index(LOANS_CONTEXT_GID_COLUMN)

    members: dict[Any, list[str]] = defaultdict(list)
    for row in rows:
        if loan_idx >= len(row) or gid_idx >= len(row):
            continue
        lid = row[loan_idx]
        gid = row[gid_idx]
        if lid is None or gid is None:
            continue
        members[gid].append(str(lid))

    if not members:
        return _EMPTY_CONTEXT

    gid_to_min: dict[Any, str] = {gid: min(loans) for gid, loans in members.items()}
    sorted_gids = sorted(gid_to_min.items(), key=lambda kv: kv[1])
    gid_to_canonical: dict[Any, int] = {gid: i + 1 for i, (gid, _) in enumerate(sorted_gids)}

    return WorkbookContext(
        gid_to_canonical=gid_to_canonical,
        gid_to_min_loan_id=gid_to_min,
    )


def _sort_rows_by_loan_canonical_key(
    rows: list[tuple[Any, ...]],
    header: tuple[str, ...],
    ctx: WorkbookContext,
) -> list[tuple[Any, ...]]:
    """Sort rows by the lex-smallest calc_loan_id in their group.

    Rows whose `collateral_group_id` is not present in the context's
    map (corruption / orphan rows) sort to the end, ordered by the
    string representation of their gid value for stability. Rows
    missing the gid column entirely sort to the very end. Pure
    set-membership ordering - no output column is used as a sort key,
    so a parity bug in `loans` / `collaterals` / `is_full_set` /
    `structure_type` cannot misalign rows.
    """
    if LOANS_CONTEXT_GID_COLUMN not in header:
        return rows
    gid_idx = header.index(LOANS_CONTEXT_GID_COLUMN)

    # Tier (smaller tier sorts first):
    #   0 = gid known in context (sort by min_loan_id)
    #   1 = gid present but unknown to context (sort by str(gid))
    #   2 = gid missing entirely
    def sort_key(row: tuple[Any, ...]) -> tuple[int, str]:
        if gid_idx >= len(row):
            return (2, "")
        gid = row[gid_idx]
        if gid is None:
            return (2, "")
        min_loan = ctx.gid_to_min_loan_id.get(gid)
        if min_loan is None:
            return (1, str(gid))
        return (0, min_loan)

    return sorted(rows, key=sort_key)


def _build_group_relabel(
    header: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> dict[Any, Any]:
    """Build a {original_group_id -> canonical_group_id} relabel map.

    Canonical id = rank of the sorted-min loan id within the group. This
    makes group ids comparable between R (igraph) and Python (networkx),
    which produce different integer labels for the same partition.
    """
    group_col = next((c for c in header if c in GROUP_ID_COLUMNS), None)
    if group_col is None:
        return {}
    group_idx = header.index(group_col)

    loan_col = next((c for c in NODE_ID_COLUMNS_BY_PRIORITY if c in header), None)
    if loan_col is None:
        return {}
    loan_idx = header.index(loan_col)

    members: dict[Any, list[Any]] = defaultdict(list)
    for row in rows:
        if group_idx >= len(row) or loan_idx >= len(row):
            continue
        gid = row[group_idx]
        lid = row[loan_idx]
        if gid is None or lid is None:
            continue
        members[gid].append(lid)

    if not members:
        return {}

    # Canonicalise: sort each group's members, then sort groups by their
    # min member, assign 1..N in that order.
    sorted_groups = sorted(members.items(), key=lambda kv: min(str(x) for x in kv[1]))
    return {gid: i + 1 for i, (gid, _) in enumerate(sorted_groups)}


def _short_repr(v: Any, limit: int = 40) -> str:
    s = repr(v)
    return s if len(s) <= limit else s[: limit - 3] + "..."


def _col_letter(idx: int) -> str:
    """1-indexed column number to Excel A/B/...AA/AB letter."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def format_report(report: DiffReport, *, max_diffs_shown: int = 30) -> str:
    """Human-readable report for the CLI tool and pytest failure messages."""
    lines: list[str] = []
    status = "PASS" if report.passed else "FAIL"
    lines.append(f"=== Parity check: {status} ===")
    lines.append(f"  actual:   {report.actual_path}")
    lines.append(f"  expected: {report.expected_path}")
    if report.actual_sheet_order != report.expected_sheet_order:
        lines.append("  SHEET-ORDER MISMATCH:")
        lines.append(f"    actual:   {list(report.actual_sheet_order)}")
        lines.append(f"    expected: {list(report.expected_sheet_order)}")

    for sd in report.sheet_diffs:
        head = f"--- {sd.sheet} ---"
        lines.append(head)
        if not sd.actual_present:
            lines.append("  MISSING IN ACTUAL")
        if not sd.expected_present:
            lines.append("  MISSING IN EXPECTED")
        if sd.actual_present and sd.expected_present:
            lines.append(
                f"  rows: actual={sd.actual_dimensions[0]} "
                f"expected={sd.expected_dimensions[0]}; "
                f"cols: actual={sd.actual_dimensions[1]} "
                f"expected={sd.expected_dimensions[1]}"
            )
            if not sd.column_order_match:
                missing = [
                    c for c in sd.column_order_expected if c not in sd.column_order_actual
                ]
                extra = [
                    c for c in sd.column_order_actual if c not in sd.column_order_expected
                ]
                lines.append("  COLUMN ORDER MISMATCH")
                if missing:
                    lines.append(f"    missing: {missing[:10]}{' ...' if len(missing) > 10 else ''}")
                if extra:
                    lines.append(f"    extra:   {extra[:10]}{' ...' if len(extra) > 10 else ''}")
        for note in sd.notes:
            lines.append(f"  note: {note}")
        if sd.cell_diffs:
            lines.append(f"  {len(sd.cell_diffs)} cell diff(s):")
            for d in sd.cell_diffs[:max_diffs_shown]:
                lines.append(d.short())
            if len(sd.cell_diffs) > max_diffs_shown:
                lines.append(f"  ... ({len(sd.cell_diffs) - max_diffs_shown} more)")
        if sd.passed and sd.actual_present and sd.expected_present:
            lines.append("  PASS")

    return "\n".join(lines)
