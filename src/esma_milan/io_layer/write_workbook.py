"""10-sheet output workbook writer (openpyxl write-only mode).

Mirrors the workbook layout in r_reference/R/pipeline.R:903-933 - ten
sheets in canonical order, each empty-or-populated according to which
pipeline stages have produced data so far.

Cell-level styling (bold + #D9D9D9 fill on header, left-align on data,
autofilter, column widths = 25) is omitted while stages stack up, since
the parity diff library reads values-only and compares values, not
formatting. Full styling lands with the Stage-10 writer.

Date encoding (Stage 6.5):
  R's openxlsx writes R Date columns as bare numeric Excel serials with
  no number-format applied. openpyxl reads such cells as `int` (no
  fractional component) rather than as `datetime.date`. To round-trip
  cell-by-cell with the staged R fixture, this writer converts Polars
  `pl.Date` columns to Excel serial integers (`(d - 1899-12-30).days`)
  before handing rows to openpyxl. This applies to every sheet that
  carries dates outside the MILAN sheet; the MILAN sheet writes ISO
  date strings in Stage 9.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import polars as pl

from esma_milan.config import OUTPUT_SHEET_ORDER

# Excel 1900-system epoch with leap-year-bug compensation. Same constant
# Stage 1's parse_iso_or_excel_date uses for the inverse direction.
_EXCEL_EPOCH: date = date(1899, 12, 30)


def _date_to_excel_serial(d: date) -> int:
    """Convert a `datetime.date` to its Excel-1900 serial integer.

    Inverse of `parse_iso_or_excel_date` in Stage 1: for every Date
    parsed by that function, this round-trips back to the same serial
    (verified by tests/unit/test_write_workbook.py).
    """
    return (d - _EXCEL_EPOCH).days


def write_pipeline_workbook(
    path: Path,
    *,
    populated_sheets: dict[str, pl.DataFrame] | None = None,
) -> None:
    """Write a 10-sheet workbook in canonical order.

    Each entry in `populated_sheets` (sheet_name -> Polars DataFrame)
    contributes a header row followed by one row per DataFrame row.
    Sheet names not present in `populated_sheets` are created empty so
    the parity harness's sheet-order check still passes.

    Polars `pl.Date` columns are encoded as Excel serial integers (see
    module docstring). All other dtypes pass through openpyxl's
    native serialisation (Int64 -> int, Float64 -> float, Boolean ->
    bool, String -> str, null -> None).
    """
    wb = openpyxl.Workbook(write_only=True)
    populated = populated_sheets or {}

    for sheet_name in OUTPUT_SHEET_ORDER:
        ws = wb.create_sheet(sheet_name)
        df = populated.get(sheet_name)
        if df is None:
            continue
        ws.append(list(df.columns))
        # Identify Date columns once per sheet so we can encode them
        # row-by-row without re-checking the schema per cell.
        date_indices = {
            i
            for i, col in enumerate(df.columns)
            if df.schema[col] == pl.Date
        }
        if not date_indices:
            for row in df.iter_rows():
                ws.append(list(row))
        else:
            for row in df.iter_rows():
                ws.append(_encode_dates(row, date_indices))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _encode_dates(
    row: tuple[Any, ...],
    date_indices: set[int],
) -> list[Any]:
    """Return a row with `pl.Date` cells converted to Excel serials.

    Null Date cells stay None (openpyxl writes them as empty cells,
    which openxlsx also produces for R `NA` Dates).
    """
    out: list[Any] = []
    for i, v in enumerate(row):
        if i in date_indices and v is not None:
            assert isinstance(v, date), (
                f"Polars pl.Date column produced non-date value {v!r}"
            )
            out.append(_date_to_excel_serial(v))
        else:
            out.append(v)
    return out


# Compatibility shim: the original API took no populated-sheets dict.
# Removed once nothing imports it.
def write_stub_workbook(path: Path) -> None:
    """Deprecated: call write_pipeline_workbook(path) instead."""
    write_pipeline_workbook(path)
