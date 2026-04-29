"""ESMA taxonomy loader.

Mirrors the taxonomy read in r_reference/R/pipeline.R:134 and the use
inside r_reference/R/utils.R:86-89:

    name_map <- taxonomy |>
      dplyr::filter(!is.na(`FIELD CODE`), !is.na(`FIELD NAME`)) |>
      dplyr::select(`FIELD CODE`, `FIELD NAME`) |>
      tibble::deframe()

Reads sheet 1 only. Filters rows where either FIELD CODE or FIELD NAME is
empty/None. Returns a {field_code -> field_name} dict that
read_csv.read_and_clean() consumes when renaming source columns.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

FIELD_CODE_COLUMN: str = "FIELD CODE"
FIELD_NAME_COLUMN: str = "FIELD NAME"


def load_taxonomy(path: Path) -> dict[str, str]:
    """Read an ESMA taxonomy XLSX and return a {field_code: field_name} dict.

    Raises:
        FileNotFoundError: if `path` does not exist.
        ValueError: if either FIELD CODE or FIELD NAME is missing from sheet 1.
    """
    if not path.exists():
        raise FileNotFoundError(f"Taxonomy file does not exist: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        ws.reset_dimensions()  # force re-scan (some writers cache 1x1)

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Taxonomy sheet 1 in {path} is empty")

        header_strs = [_norm(h) for h in header]
        try:
            code_idx = header_strs.index(FIELD_CODE_COLUMN)
        except ValueError:
            raise ValueError(
                f"Taxonomy must contain {FIELD_CODE_COLUMN!r} column. "
                f"Found columns: {[h for h in header_strs if h]}"
            ) from None
        try:
            name_idx = header_strs.index(FIELD_NAME_COLUMN)
        except ValueError:
            raise ValueError(
                f"Taxonomy must contain {FIELD_NAME_COLUMN!r} column. "
                f"Found columns: {[h for h in header_strs if h]}"
            ) from None

        mapping: dict[str, str] = {}
        for row in rows_iter:
            code = _cell(row, code_idx)
            name = _cell(row, name_idx)
            if code is None or name is None:
                continue
            # Mirror R's deframe(): later entries overwrite earlier ones with
            # the same key. The synthetic taxonomy doesn't exercise this; the
            # behaviour is documented to match R exactly should it ever arise.
            mapping[code] = name
        return mapping
    finally:
        wb.close()


def _norm(value: object) -> str:
    """Normalise a header cell to a stripped string. None -> ''."""
    if value is None:
        return ""
    return str(value).strip()


def _cell(row: tuple[object, ...], idx: int) -> str | None:
    """Return a stripped string for a data cell, or None if missing/empty."""
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None
